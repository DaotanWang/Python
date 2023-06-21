import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active #默认创建了一个sheet，active选中这个sheet进行操作

#创建多个sheet
#sheet1 = wb.create_sheet('Sheet1')  
#sheet2 = wb.create_sheet('Sheet2')

sheet['B1'] = 'ChannelID' #通过单元格塞进元素
sheet1.cell(2, 1, 'World') # B2单元格写入'World' 通过行列位置塞进元素

#循环塞进元素：

for c in range(0,len(df1_list))

  cell1 = 'A' + str(c + 2) #c是数字要变成str
  sheet[cell1] = str(text)

#表格保存
wb.save(r'C:\data.xlsx')
